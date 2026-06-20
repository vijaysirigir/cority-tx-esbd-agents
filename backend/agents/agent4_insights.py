"""
Agent 4 - Search Insights.

Rolls up the whole workbook for the current search into a one-glance summary and
writes it to sheet 4. In addition to the headline counts (Raw Results, ESHQ
Opportunities, RFPs Detailed, Attachments Saved) it reports the sales-readiness
metrics pulled from the detailed RFPs:

  * Total Solicitation IDs        (from the raw search)
  * Total Contact Names           (from detailed RFPs)
  * Total Contact Numbers         (from detailed RFPs)
  * Total Contact Emails          (from detailed RFPs)
  * Sum of Estimated Value        (from detailed RFPs, where published)

Contact + value metrics come from sheet 3, so run Agent 3 on the opportunities
you care about first; otherwise those tallies are zero.

Run standalone:
    echo '{}' | python agents/agent4_insights.py --out r.json
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import config
from agents import browser as b
from agents import excel_store
from openpyxl import load_workbook


def _read_sheet(wb, name) -> tuple[list[str], list[dict]]:
    if name not in wb.sheetnames:
        return [], []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = [dict(zip(headers, r)) for r in rows[1:]]
    return headers, out


def _col(headers: list[str], *hints: str) -> str | None:
    for h in headers:
        hl = h.lower()
        if any(hint in hl for hint in hints):
            return h
    return None


def _nonempty(rows: list[dict], col: str | None) -> int:
    if not col:
        return 0
    return sum(1 for r in rows if str(r.get(col, "") or "").strip())


def _money(s) -> float:
    if s is None:
        return 0.0
    s = str(s)
    m = re.search(r"[\d][\d,]*(?:\.\d+)?", s.replace("$", ""))
    if not m:
        return 0.0
    try:
        v = float(m.group(0).replace(",", ""))
    except ValueError:
        return 0.0
    low = s.lower()
    if "million" in low or re.search(r"\dm\b", low):
        v *= 1_000_000
    elif "billion" in low:
        v *= 1_000_000_000
    elif "thousand" in low or re.search(r"\dk\b", low):
        v *= 1_000
    return v


def run(params: dict) -> dict:
    result: dict = {"ok": False, "agent": "agent4"}
    if not config.MASTER_WORKBOOK.exists():
        result["error"] = ("No workbook yet. Run Agent 1 (and ideally 2 & 3) "
                           "first, then generate insights.")
        return result

    b.log("Agent 4 - summarising the current search...")
    wb = load_workbook(config.MASTER_WORKBOOK, read_only=True, data_only=True)

    raw_h, raw_rows = _read_sheet(wb, config.SHEET_RAW)
    opp_h, opp_rows = _read_sheet(wb, config.SHEET_OPPS)
    det_h, det_rows = _read_sheet(wb, config.SHEET_DETAILS)

    # --- Total solicitation IDs (the search) ---
    sid_col = _col(raw_h, "solicitation id", "solicitation", "id")
    if sid_col:
        ids = {str(r.get(sid_col, "")).strip() for r in raw_rows
               if str(r.get(sid_col, "") or "").strip()}
        total_ids = len(ids)
    else:
        total_ids = len(raw_rows)
    b.log(f"  - {total_ids} solicitation IDs in the raw search")

    # --- Contact + value metrics (from detailed RFPs) ---
    if not det_rows:
        b.log("  - no RFP Details sheet yet; contact/value metrics are 0 "
              "(run Agent 3 to populate them)")
    name_col = _col(det_h, "contact name")
    num_col = _col(det_h, "contact number", "contact phone")
    email_col = _col(det_h, "contact email")
    val_col = _col(det_h, "estimated value", "contract value", "estimated amount")

    contact_names = _nonempty(det_rows, name_col)
    contact_numbers = _nonempty(det_rows, num_col)
    contact_emails = _nonempty(det_rows, email_col)

    value_sum = 0.0
    value_count = 0
    if val_col:
        for r in det_rows:
            v = _money(r.get(val_col, ""))
            if v > 0:
                value_sum += v
                value_count += 1
    b.log(f"  - contacts: {contact_names} names, {contact_numbers} numbers, "
          f"{contact_emails} emails")
    b.log(f"  - estimated value: ${value_sum:,.0f} across {value_count} RFP(s)")

    # --- Opportunity tier breakdown (from sheet 2) ---
    action_col = _col(opp_h, "action")
    tiers: dict[str, int] = {}
    if action_col:
        for r in opp_rows:
            a = str(r.get(action_col, "") or "").strip()
            if a:
                tiers[a] = tiers.get(a, 0) + 1
    score_col = _col(opp_h, "opportunity score", "score")
    scores = [float(r[score_col]) for r in opp_rows
              if score_col and isinstance(r.get(score_col), (int, float))]
    top_score = max(scores) if scores else 0
    avg_score = round(sum(scores) / len(scores)) if scores else 0

    # --- Attachments saved (actual files on disk) ---
    attachments_saved = sum(1 for p in config.ATTACHMENTS_DIR.rglob("*")
                            if p.is_file() and not p.name.startswith("_details"))

    value_display = f"${value_sum:,.0f}" if value_sum else "Not published"

    metrics = {
        "solicitation_ids": total_ids,
        "eshq_opportunities": len(opp_rows),
        "rfps_detailed": len(det_rows),
        "attachments_saved": attachments_saved,
        "contact_names": contact_names,
        "contact_numbers": contact_numbers,
        "contact_emails": contact_emails,
        "estimated_value_sum": round(value_sum, 2),
        "estimated_value_display": value_display,
        "rfps_with_value": value_count,
        "top_score": top_score,
        "avg_score": avg_score,
    }

    # --- Write sheet 4 ---
    ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    rows = [
        ["Search summarised", ts],
        ["Total Solicitation IDs (search)", total_ids],
        ["ESHQ Opportunities (kept)", len(opp_rows)],
        ["RFPs Detailed", len(det_rows)],
        ["Attachments Saved", attachments_saved],
        ["", ""],
        ["Total Contact Names", contact_names],
        ["Total Contact Numbers", contact_numbers],
        ["Total Contact Emails", contact_emails],
        ["Sum of Estimated Value", value_display],
        ["RFPs with a published value", value_count],
        ["", ""],
        ["Top opportunity score", top_score],
        ["Average opportunity score", avg_score],
    ]
    for action in ("Immediate Executive Review", "High Priority Opportunity",
                   "Sales Review", "Monitor"):
        if tiers.get(action):
            rows.append([f"  - {action}", tiers[action]])

    excel_store.write_sheet(config.SHEET_INSIGHTS, ["Metric", "Value"], rows,
                            index=3, wide_cols=[1])
    b.log(f"  - wrote summary to '{config.SHEET_INSIGHTS}'")

    result.update(ok=True, workbook=str(config.MASTER_WORKBOOK),
                  metrics=metrics, tier_breakdown=tiers, generated_at=ts)
    return result


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out")
    args = ap.parse_args()
    params = b.read_params_from_stdin()
    res = run(params)
    b.emit_result(res, args.out)
