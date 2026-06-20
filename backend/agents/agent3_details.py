"""
Agent 3 - RFP Detail & Attachment Extraction.

For each solicitation ID (passed in, or read from the ESHQ Opportunities sheet
produced by Agent 2), this agent:

  1. searches the ESBD for that ID,
  2. clicks the solicitation title hyperlink to open the detail page,
  3. extracts every detail it can find (header fields, label/value pairs,
     description, contact, dates), and
  4. downloads all attachments into data/attachments/<solicitation id>/,

then writes one row per solicitation into sheet 3 of the master workbook, with
the full field set captured and the saved attachment paths referenced.

Run standalone:
    echo '{"solicitation_ids":["...id..."]}' | python agents/agent3_details.py --out r.json
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from urllib.parse import urljoin

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import config
from agents import browser as b
from agents import excel_store
from openpyxl import load_workbook

DOC_EXT = (".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".zip", ".ppt",
           ".pptx", ".txt", ".rtf", ".jpg", ".png", ".tif", ".tiff")
ATTACH_HINTS = ("attachment", "download", "blobfile", "/file", "document",
                "getfile", "viewfile")


def _ids_from_opps_sheet(limit: int | None) -> list[str]:
    if not config.MASTER_WORKBOOK.exists():
        return []
    wb = load_workbook(config.MASTER_WORKBOOK, read_only=True)
    if config.SHEET_OPPS not in wb.sheetnames:
        return []
    ws = wb[config.SHEET_OPPS]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h) for h in rows[0]]
    try:
        idx = header.index("Solicitation ID")
    except ValueError:
        idx = 1
    ids = []
    for r in rows[1:]:
        if idx < len(r) and r[idx]:
            ids.append(str(r[idx]).strip())
    if limit:
        ids = ids[:limit]
    return ids


def run(params: dict) -> dict:
    ids = params.get("solicitation_ids") or []
    if isinstance(ids, str):
        ids = [ids]
    ids = [str(i).strip() for i in ids if str(i).strip()]
    headful = bool(params.get("headful"))
    limit = params.get("limit")

    if not ids:
        ids = _ids_from_opps_sheet(limit)
        if ids:
            b.log(f"Agent 3 - pulled {len(ids)} IDs from '{config.SHEET_OPPS}'")

    result: dict = {"ok": False, "agent": "agent3"}
    if not ids:
        result["error"] = (
            "No solicitation IDs provided and none found in the ESHQ "
            "Opportunities sheet. Run Agent 2 first or pass solicitation_ids."
        )
        return result

    records = []
    errors = []

    try:
        with b.browser_page(headful=headful) as page:
            for i, sid in enumerate(ids, 1):
                b.log(f"[{i}/{len(ids)}] Looking up solicitation {sid}...")
                try:
                    rec = _extract_one(page, sid)
                    records.append(rec)
                    b.log(f"  - {rec['title'][:60]!r} - "
                          f"{rec['attachment_count']} attachment(s)")
                except Exception as e:  # noqa: BLE001
                    b.log(f"  !! failed on {sid}: {e}")
                    dbg = b.dump_debug(page, f"agent3_{_safe(sid)}")
                    errors.append({"id": sid, "error": str(e), "debug": dbg})
                    records.append(_empty_record(sid, error=str(e)))
    except Exception as e:  # noqa: BLE001
        result["error"] = str(e)
        return result

    # Write sheet 3
    out_headers = [
        "Solicitation ID", "Title", "Agency", "Status", "Posted Date",
        "Due / Close Date", "Contact", "Detail URL", "Description",
        "Key Fields", "# Attachments", "Attachment Files", "Saved Folder",
    ]
    out_rows = [[
        r["solicitation_id"], r["title"], r["agency"], r["status"],
        r["posted"], r["due"], r["contact"], r["url"], r["description"],
        r["key_fields_text"], r["attachment_count"],
        "\n".join(r["attachment_files"]), r["folder"],
    ] for r in records]

    wb = excel_store.write_sheet(
        config.SHEET_DETAILS, out_headers, out_rows, index=2,
        wide_cols=[2, 9, 10, 12],  # Title, Description, Key Fields, Attachment Files
    )
    b.log(f"Agent 3 - wrote {len(out_rows)} rows to '{config.SHEET_DETAILS}'")

    result.update(
        ok=True, workbook=str(wb), processed=len(records),
        errors=errors, records=[_record_summary(r) for r in records],
    )
    return result


# ---------------------------------------------------------------------------
def _extract_one(page, sid: str) -> dict:
    # 1. Go straight to the detail page — /esbd/<id> works for both the readable
    #    and internal ID formats the ESBD uses.
    page.goto(config.DETAIL_URL_TEMPLATE.format(id=sid), wait_until="domcontentloaded")
    page.wait_for_timeout(2500)

    # 2. If that wasn't a detail page, fall back to searching by ID + clicking.
    if not _is_detail_page(page):
        b.log("  - direct detail not found; searching by ID instead")
        page.goto(config.ESBD_URL, wait_until="domcontentloaded")
        page.wait_for_timeout(1200)
        b.safe_fill(page, "solicitation_id", sid)
        if not b.safe_click(page, "search_btn"):
            page.keyboard.press("Enter")
        page.wait_for_timeout(2500)
        link = b.first_locator(page, config.SELECTORS["result_title_link"], timeout=8000)
        if not link:
            raise RuntimeError("no detail page or matching result found for this ID")
        try:
            link.click()
        except Exception:
            href = link.get_attribute("href")
            if href:
                page.goto(urljoin(config.BASE_URL, href))
        page.wait_for_timeout(2500)
    detail_url = page.url

    # Make sure the Attachments tab's content is rendered.
    _open_attachments_tab(page)

    # 3. extract details
    body = _main_text(page)
    fields = _extract_fields(page, body)
    title = _heading_title(page, body) or _pick(fields, ["title"]) or sid
    rec = {
        "solicitation_id": _pick(fields, ["solicitation id"]) or sid,
        "title": title,
        "agency": _pick(fields, ["agency", "member name", "organization",
                                 "customer"]),
        "status": _pick(fields, ["status"]),
        "posted": _pick(fields, ["posting date", "posted", "issue date"]),
        "due": _pick(fields, ["response due", "due date", "close", "deadline"]),
        "contact": _contact(fields),
        "url": detail_url,
        "description": (_section_after(body, ["Solicitation Description",
                                              "Description", "Scope"])
                        or _pick(fields, ["description", "summary", "scope"])
                        or body[:4000]),
        "fields": fields,
        "key_fields_text": _fields_to_text(fields),
    }

    # 4. attachments
    folder = config.ATTACHMENTS_DIR / _safe(sid)
    files = _download_attachments(page, folder)
    rec["folder"] = str(folder) if files else ""
    rec["attachment_files"] = [Path(f).name for f in files]
    rec["attachment_paths"] = files
    rec["attachment_count"] = len(files)

    # Save the full extracted text/fields next to the attachments for reference.
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "_details.json").write_text(
        json.dumps(rec, default=str, indent=2), encoding="utf-8")
    return rec


def _extract_fields(page, body: str | None = None) -> dict:
    """Parse the ESBD detail layout into label -> value pairs.

    ESBD renders each field as a label line ending in ':' followed (after blank
    lines) by the value on the next non-empty line, e.g.::

        Solicitation ID:
        0008984
        Status:
        Posted
    """
    fields: dict[str, str] = {}
    text = body if body is not None else _main_text(page)
    lines = [ln.replace("\xa0", " ").strip() for ln in text.splitlines()]
    n = len(lines)

    label_re = re.compile(r"^(.{2,60}?):$")
    for i, line in enumerate(lines):
        # inline "Label: value"
        mm = re.match(r"^([A-Za-z][A-Za-z /&#]{1,40}):\s+(.{1,400})$", line)
        if mm:
            fields.setdefault(mm.group(1).strip(), mm.group(2).strip())
            continue
        # label on its own line, value on the next non-empty line
        m = label_re.match(line)
        if m:
            label = m.group(1).strip()
            j = i + 1
            while j < n and not lines[j]:
                j += 1
            if j < n and not label_re.match(lines[j]):
                fields.setdefault(label, lines[j])
    return fields


def _is_detail_page(page) -> bool:
    """A detail page has the 'Download PDF' button or a Solicitation ID label."""
    try:
        if page.locator(config.SELECTORS["download_pdf_btn"][0]).count() > 0:
            return True
    except Exception:
        pass
    body = _main_text(page)
    return ("Solicitation ID" in body) and ("Page 1 of" not in body)


def _open_attachments_tab(page) -> None:
    loc = b.first_locator(page, config.SELECTORS.get("attachments_tab", []),
                          timeout=2500)
    if loc:
        try:
            loc.click()
            page.wait_for_timeout(800)
        except Exception:
            pass


def _heading_title(page, body: str = "") -> str:
    t = _heading(page)
    if t and t.lower() not in ("esbd", "electronic state business daily"):
        return t
    # Fallback: the name sits just before "Follow Solicitation" / "Solicitation ID"
    for line in (body or _main_text(page)).splitlines():
        s = line.replace("\xa0", " ").strip()
        if s and s not in ("Download PDF", "Follow Solicitation",
                           "Electronic State Business Daily") and ":" not in s \
                and len(s) > 8:
            return s
    return ""


def _section_after(body: str, labels: list[str], max_len: int = 6000) -> str:
    """Return the text following a label up to the next 'Label:' line."""
    lines = [ln.replace("\xa0", " ").rstrip() for ln in body.splitlines()]
    label_re = re.compile(r"^.{2,60}?:\s*$")
    for li, line in enumerate(lines):
        for lab in labels:
            if line.strip().lower().startswith(lab.lower()):
                out = []
                for nxt in lines[li + 1:]:
                    if label_re.match(nxt.strip()) and out:
                        break
                    if nxt.strip():
                        out.append(nxt.strip())
                    if sum(len(x) for x in out) > max_len:
                        break
                if out:
                    return "\n".join(out)[:max_len]
    return ""


def _contact(fields: dict) -> str:
    parts = []
    for k in ("Contact Name", "Contact Number", "Contact Email"):
        if k in fields:
            parts.append(fields[k])
    return " | ".join(parts) if parts else _pick(
        fields, ["contact", "buyer", "email", "phone"])


def _download_attachments(page, folder: Path) -> list[str]:
    """Save all attachments. ESBD attachment links carry the real file URL in a
    `data-href` attribute (action=downloadURL); fall back to scanning regular
    hrefs for anything that looks like a document."""
    saved: list[str] = []
    seen: set[str] = set()
    candidates: list[tuple[str, str]] = []  # (full_url, filename_text)

    # Primary: data-href anchors.
    try:
        locs = page.locator("a[data-href], a[data-action='downloadURL']")
        for i in range(locs.count()):
            a = locs.nth(i)
            dh = a.get_attribute("data-href")
            if not dh:
                continue
            text = (a.inner_text() or "").strip()
            full = urljoin(page.url, dh)
            if full not in seen:
                seen.add(full)
                candidates.append((full, text))
    except Exception:
        pass

    # Secondary: plain document hrefs.
    try:
        anchors = page.locator("a[href]")
        for i in range(min(anchors.count(), 400)):
            href = anchors.nth(i).get_attribute("href") or ""
            low = href.lower()
            if low.endswith(DOC_EXT) or any(h in low for h in ATTACH_HINTS):
                full = urljoin(page.url, href)
                if full not in seen:
                    seen.add(full)
                    candidates.append((full, (anchors.nth(i).inner_text() or "").strip()))
    except Exception:
        pass

    if not candidates:
        return saved

    folder.mkdir(parents=True, exist_ok=True)
    for idx, (href, text) in enumerate(candidates, 1):
        name = _filename_from(href, text, idx)
        target = folder / name
        if _fetch_to(page, href, target):
            saved.append(str(target))
            b.log(f"     v {name}")
        elif _click_download(page, href, target):
            saved.append(str(target))
            b.log(f"     v {name} (via click)")
        else:
            b.log(f"     x could not download {name}")

    # Best-effort: also grab the generated 'Download PDF' solicitation summary.
    try:
        btn = b.first_locator(page, config.SELECTORS.get("download_pdf_btn", []),
                              timeout=2000)
        if btn:
            target = folder / "_Solicitation_Summary.pdf"
            with page.expect_download(timeout=30000) as dl:
                btn.click()
            dl.value.save_as(str(target))
            if target.exists():
                saved.append(str(target))
                b.log("     v _Solicitation_Summary.pdf")
    except Exception:
        pass

    return saved


def _fetch_to(page, href: str, target: Path) -> bool:
    try:
        resp = page.context.request.get(href, timeout=config.DOWNLOAD_TIMEOUT_MS)
        if not resp.ok:
            return False
        ctype = (resp.headers or {}).get("content-type", "")
        body = resp.body()
        # Skip if it's clearly an HTML page rather than a file.
        if "text/html" in ctype and not target.suffix.lower() in (".html", ".htm"):
            if len(body) < 100 or b"<html" in body[:1000].lower():
                return False
        target.write_bytes(body)
        return target.stat().st_size > 0
    except Exception:
        return False


def _click_download(page, href: str, target: Path) -> bool:
    # `href` may be a normal href or a resolved data-href; the data-href stored
    # on the anchor is a relative path, so match on its tail.
    tail = href.split("txsmartbuy.gov", 1)[-1]
    for sel in (f"a[href='{href}']", f"a[data-href='{tail}']",
                f"a[data-href*='{tail.split('?')[0]}']"):
        try:
            loc = page.locator(sel).first
            with page.expect_download(timeout=config.DOWNLOAD_TIMEOUT_MS) as dl:
                loc.click()
            dl.value.save_as(str(target))
            if target.exists():
                return True
        except Exception:
            continue
    return False


# ---------------------------------------------------------------------------
def _main_text(page) -> str:
    for sel in ("main", "#content", ".content", "article", "body"):
        try:
            loc = page.locator(sel).first
            if loc.count():
                t = loc.inner_text()
                if t and len(t) > 80:
                    return t
        except Exception:
            continue
    try:
        return page.inner_text("body")
    except Exception:
        return ""


def _heading(page) -> str:
    for sel in ("h1", "h2", ".page-title", "title"):
        try:
            loc = page.locator(sel).first
            if loc.count():
                t = (loc.inner_text() or "").strip()
                if t:
                    return t
        except Exception:
            continue
    return ""


def _pick(fields: dict, keys: list[str]) -> str:
    for want in keys:
        for k, v in fields.items():
            if want in k.lower():
                return v
    return ""


def _fields_to_text(fields: dict) -> str:
    return "\n".join(f"{k}: {v}" for k, v in fields.items())


def _filename_from(href: str, text: str, idx: int) -> str:
    # The link text is usually the real filename (e.g. ESBD_File_..._RFO.pdf).
    if text and re.search(r"\.[A-Za-z0-9]{2,5}$", text):
        return _safe(text)[:140]
    # Otherwise derive the extension from the media endpoint's _xt param.
    m = re.search(r"_xt=(\.[A-Za-z0-9]+)", href)
    ext = m.group(1) if m else ""
    base = href.split("?")[0].rstrip("/").split("/")[-1]
    if base and "." in base and base.lower() != "media.nl" and len(base) < 120:
        return _safe(base)
    if text:
        nm = _safe(text)[:120]
        return nm + ext if ext and not nm.lower().endswith(ext.lower()) else nm
    return f"attachment_{idx}{ext or '.pdf'}"


def _safe(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_") or "file"


def _empty_record(sid: str, error: str = "") -> dict:
    return {
        "solicitation_id": sid, "title": f"(not found) {sid}", "agency": "",
        "status": "", "posted": "", "due": "", "contact": "", "url": "",
        "description": error, "fields": {}, "key_fields_text": "",
        "folder": "", "attachment_files": [], "attachment_paths": [],
        "attachment_count": 0,
    }


def _record_summary(r: dict) -> dict:
    return {
        "solicitation_id": r["solicitation_id"], "title": r["title"],
        "agency": r["agency"], "status": r["status"], "due": r["due"],
        "url": r["url"], "attachment_count": r["attachment_count"],
        "attachment_files": r["attachment_files"],
    }


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out")
    args = ap.parse_args()
    params = b.read_params_from_stdin()
    res = run(params)
    b.emit_result(res, args.out)
