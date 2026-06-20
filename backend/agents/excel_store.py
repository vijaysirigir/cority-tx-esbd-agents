"""
Shared Excel workbook helper.

All three agents read from / write to a single workbook
(config.MASTER_WORKBOOK) with three sheets:

    1. Raw Search Results   <- Agent 1 (the exported CSV, loaded verbatim)
    2. ESHQ Opportunities   <- Agent 2 (Cority-relevant rows + scoring)
    3. RFP Details          <- Agent 3 (per-solicitation detail + attachments)

Sheets are written with Cority-styled headers (navy fill, white bold text),
a frozen header row, an auto-filter and sensible column widths.
"""
from __future__ import annotations

import csv
import datetime as _dt
import os
import zipfile
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config

# Cority palette (no leading '#': openpyxl wants 'AARRGGBB' or 'RRGGBB')
NAVY = "002448"
ORANGE = "D2640A"
LIGHT = "F4F6F9"

_HEADER_FILL = PatternFill("solid", fgColor=NAVY)
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
_ZEBRA_FILL = PatternFill("solid", fgColor=LIGHT)
_THIN = Side(style="thin", color="E2E7EE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="top", wrap_text=True)
_TOP = Alignment(vertical="top")


def _new_book() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet - we name our own
    return wb


def _open_book() -> Workbook:
    """Open the master workbook, self-healing if it is missing or corrupt.

    OneDrive sync (the data dir lives under OneDrive) can occasionally leave a
    partially written file. Rather than crash, we set the bad file aside and
    start fresh so the pipeline keeps working (re-run earlier agents to refill).
    """
    if not config.MASTER_WORKBOOK.exists():
        return _new_book()
    try:
        return load_workbook(config.MASTER_WORKBOOK)
    except (zipfile.BadZipFile, KeyError, OSError, Exception) as e:  # noqa: BLE001
        bad = config.MASTER_WORKBOOK.with_suffix(
            f".corrupt_{_dt.datetime.now():%Y%m%d_%H%M%S}.xlsx")
        try:
            config.MASTER_WORKBOOK.rename(bad)
            print(f"[excel_store] workbook unreadable ({e}); moved to {bad.name}, "
                  f"starting a fresh one", flush=True)
        except Exception:
            pass
        return _new_book()


def write_sheet(
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    index: int | None = None,
    wide_cols: Sequence[int] | None = None,
) -> Path:
    """Create/replace `sheet_name` with `headers` + `rows`. Returns workbook path.

    `wide_cols` is a list of 1-based column numbers that should be wide + wrapped
    (used for long free-text fields like descriptions).
    """
    wb = _open_book()

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name, index=index)

    wide = set(wide_cols or [])

    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))

    _style_sheet(ws, len(headers), wide)
    _reorder(wb)
    _atomic_save(wb)
    return config.MASTER_WORKBOOK


def _atomic_save(wb: Workbook) -> None:
    """Save to a temp file in the same folder, then atomically replace the
    target. Prevents a half-written (corrupt) workbook if anything interrupts
    the save (OneDrive sync, crash, antivirus).

    On Windows the destination can be briefly locked (OneDrive/AV/another
    reader); retry a few times, then fall back to a direct save so the write
    still lands.
    """
    import time
    config.MASTER_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    tmp = config.MASTER_WORKBOOK.with_suffix(f".tmp_{os.getpid()}.xlsx")
    wb.save(tmp)
    for attempt in range(6):
        try:
            os.replace(tmp, config.MASTER_WORKBOOK)  # atomic on same volume
            return
        except PermissionError:
            time.sleep(0.5 * (attempt + 1))
    # Last resort: direct save (non-atomic) so data isn't lost.
    try:
        wb.save(config.MASTER_WORKBOOK)
    finally:
        try:
            tmp.unlink()
        except Exception:
            pass


def _style_sheet(ws: Worksheet, ncols: int, wide: set[int]) -> None:
    # Header row
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _BORDER

    # Body
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER
            cell.alignment = _WRAP if c in wide else _TOP
            if r % 2 == 0:
                cell.fill = _ZEBRA_FILL

    # Column widths - measure content, clamp.
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        if c in wide:
            ws.column_dimensions[letter].width = 60
            continue
        longest = 10
        for r in range(1, min(max_row, 200) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                longest = max(longest, min(len(str(v)), 48))
        ws.column_dimensions[letter].width = longest + 3

    ws.freeze_panes = "A2"
    if max_row >= 1 and ncols >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max_row}"
    ws.row_dimensions[1].height = 26


_SHEET_ORDER = [config.SHEET_RAW, config.SHEET_OPPS, config.SHEET_DETAILS]


def _reorder(wb: Workbook) -> None:
    """Keep the three known sheets in 1-2-3 order."""
    desired = [s for s in _SHEET_ORDER if s in wb.sheetnames]
    others = [s for s in wb.sheetnames if s not in _SHEET_ORDER]
    wb._sheets.sort(key=lambda s: (desired + others).index(s.title))  # type: ignore[attr-defined]


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------
def read_csv(path: str | Path) -> tuple[list[str], list[dict]]:
    """Read a CSV into (headers, list-of-dict-rows). Tolerant of BOM/encoding."""
    p = Path(path)
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with p.open("r", encoding=enc, newline="") as fh:
                reader = csv.DictReader(fh)
                headers = reader.fieldnames or []
                rows = [dict(r) for r in reader]
            return [h for h in headers if h is not None], rows
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"Could not decode {p}")


def load_csv_into_raw_sheet(csv_path: str | Path) -> tuple[Path, int, list[str]]:
    """Load an exported CSV verbatim into sheet 1. Returns (wb, nrows, headers)."""
    headers, rows = read_csv(csv_path)
    data = [[r.get(h, "") for h in headers] for r in rows]
    write_sheet(config.SHEET_RAW, headers, data, index=0)
    return config.MASTER_WORKBOOK, len(rows), headers
