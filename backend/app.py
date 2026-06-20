"""
Cority ESBD Agent Suite — Flask backend.

Serves the Cority-branded front end and orchestrates the three agents. Each
agent runs as an isolated subprocess (clean Playwright environment, crash
isolation) that streams progress to a log file; the front end polls
/api/status/<job_id> for live updates and the final JSON result.

Run:
    cd backend
    python app.py
then open http://127.0.0.1:5000
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
import threading
import uuid
from pathlib import Path

from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook

import config

BACKEND_DIR = Path(__file__).resolve().parent

AGENT_SCRIPTS = {
    "agent1": BACKEND_DIR / "agents" / "agent1_search.py",
    "agent2": BACKEND_DIR / "agents" / "agent2_filter.py",
    "agent3": BACKEND_DIR / "agents" / "agent3_details.py",
    "agent4": BACKEND_DIR / "agents" / "agent4_insights.py",
}

app = Flask(__name__, static_folder=None)
CORS(app)

# job_id -> {agent, proc, log_path, out_path}
JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.Lock()


# ---------------------------------------------------------------------------
# Front end
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return send_from_directory(config.FRONTEND_DIR, "index.html")


@app.route("/<path:path>")
def static_files(path):
    target = config.FRONTEND_DIR / path
    if target.exists():
        return send_from_directory(config.FRONTEND_DIR, path)
    # also serve brand assets (logo)
    asset = config.ASSETS_DIR / path
    if asset.exists():
        return send_from_directory(config.ASSETS_DIR, path)
    return ("Not found", 404)


@app.route("/assets/<path:path>")
def assets(path):
    return send_from_directory(config.ASSETS_DIR, path)


# ---------------------------------------------------------------------------
# Run an agent (returns a job id immediately)
# ---------------------------------------------------------------------------
@app.route("/api/run/<agent>", methods=["POST"])
def run_agent(agent):
    if agent not in AGENT_SCRIPTS:
        return jsonify({"error": f"unknown agent '{agent}'"}), 400

    params = request.get_json(silent=True) or {}
    job_id = uuid.uuid4().hex[:12]
    log_path = config.DEBUG_DIR / f"job_{job_id}.log"
    out_path = config.DEBUG_DIR / f"job_{job_id}.result.json"

    log_fh = open(log_path, "w", encoding="utf-8")
    child_env = {**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"}
    proc = subprocess.Popen(
        [sys.executable, str(AGENT_SCRIPTS[agent]), "--out", str(out_path)],
        cwd=str(BACKEND_DIR),
        stdin=subprocess.PIPE,
        stdout=log_fh,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        env=child_env,
    )
    try:
        proc.stdin.write(json.dumps(params))
        proc.stdin.close()
    except Exception:
        pass

    with JOBS_LOCK:
        JOBS[job_id] = {
            "agent": agent, "proc": proc, "log_fh": log_fh,
            "log_path": log_path, "out_path": out_path,
        }
    return jsonify({"job_id": job_id, "agent": agent})


@app.route("/api/status/<job_id>")
def status(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"error": "unknown job"}), 404

    proc = job["proc"]
    running = proc.poll() is None
    log_text = ""
    try:
        log_text = Path(job["log_path"]).read_text(encoding="utf-8", errors="ignore")
    except Exception:
        pass

    result = None
    if not running:
        try:
            job["log_fh"].close()
        except Exception:
            pass
        out = Path(job["out_path"])
        if out.exists():
            try:
                result = json.loads(out.read_text(encoding="utf-8"))
            except Exception as e:  # noqa: BLE001
                result = {"ok": False, "error": f"could not parse result: {e}"}
        else:
            # fall back to parsing the RESULT_JSON_BELOW marker from the log
            result = _parse_log_result(log_text) or {
                "ok": False,
                "error": "agent exited without a result (see log)",
            }

    # Keep only the log lines (drop the trailing raw JSON dump for display).
    display_log = _clean_log(log_text)
    return jsonify({
        "running": running,
        "agent": job["agent"],
        "log": display_log,
        "result": result,
    })


def _clean_log(text: str) -> str:
    if "RESULT_JSON_BELOW" in text:
        text = text.split("RESULT_JSON_BELOW")[0]
    return text.strip()


def _parse_log_result(text: str) -> dict | None:
    if "RESULT_JSON_BELOW" in text:
        tail = text.split("RESULT_JSON_BELOW", 1)[1].strip()
        try:
            return json.loads(tail.splitlines()[0])
        except Exception:
            return None
    return None


# ---------------------------------------------------------------------------
# Dashboard state + downloads
# ---------------------------------------------------------------------------
@app.route("/api/state")
def state():
    info = {
        "workbook_exists": config.MASTER_WORKBOOK.exists(),
        "workbook_path": str(config.MASTER_WORKBOOK),
        "sheets": {},
        "latest_csv": None,
        "attachment_count": 0,
    }
    if config.MASTER_WORKBOOK.exists():
        wb = None
        try:
            wb = load_workbook(config.MASTER_WORKBOOK, read_only=True)
            for name in wb.sheetnames:
                ws = wb[name]
                info["sheets"][name] = max(0, (ws.max_row or 1) - 1)
        except Exception:
            pass
        finally:
            # read_only mode keeps the file handle open; closing it releases the
            # lock so the agents can rewrite the workbook on Windows.
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
    csvs = sorted(config.DOWNLOADS_DIR.glob("*.csv"),
                  key=lambda p: p.stat().st_mtime, reverse=True)
    if csvs:
        info["latest_csv"] = csvs[0].name
    info["attachment_count"] = sum(
        1 for _ in config.ATTACHMENTS_DIR.rglob("*") if _.is_file())
    return jsonify(info)


@app.route("/api/download/workbook")
def download_workbook():
    if not config.MASTER_WORKBOOK.exists():
        return ("No workbook yet — run the agents first.", 404)
    return send_file(config.MASTER_WORKBOOK, as_attachment=True,
                     download_name=config.MASTER_WORKBOOK.name)


@app.route("/api/download/csv")
def download_csv():
    csvs = sorted(config.DOWNLOADS_DIR.glob("*.csv"),
                  key=lambda p: p.stat().st_mtime, reverse=True)
    if not csvs:
        return ("No CSV yet — run Agent 1 first.", 404)
    return send_file(csvs[0], as_attachment=True, download_name=csvs[0].name)


if __name__ == "__main__":
    # In the cloud (Render) the platform injects $PORT and the app must listen on
    # 0.0.0.0. Locally this still works as http://127.0.0.1:5000.
    port = int(os.environ.get("PORT", "5000"))
    print("=" * 60)
    print("  Cority ESBD Agent Suite")
    print(f"  Listening on http://0.0.0.0:{port}")
    print(f"  Workbook:    {config.MASTER_WORKBOOK}")
    print("=" * 60)
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
